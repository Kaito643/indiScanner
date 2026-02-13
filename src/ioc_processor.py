import re
import os
from typing import List, Set

class IOCProcessor:
    # Regex for SHA256: 64 hex characters
    SHA256_PATTERN = re.compile(r'\b[a-fA-F0-9]{64}\b')

    @staticmethod
    def extract_hashes_from_text(text: str) -> Set[str]:
        """Extracts unique SHA256 hashes from a string."""
        return set(IOCProcessor.SHA256_PATTERN.findall(text))

    @staticmethod
    def extract_hashes_from_file(file_path: str) -> Set[str]:
        """Reads a file (txt, pdf, xlsx) and extracts SHA256 hashes."""
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        
        ext = os.path.splitext(file_path)[1].lower()
        content = ""

        try:
            if ext == '.pdf':
                from pypdf import PdfReader
                reader = PdfReader(file_path)
                for page in reader.pages:
                    content += page.extract_text() + "\n"
            
            elif ext in ['.xlsx', '.xlsm']:
                import openpyxl
                wb = openpyxl.load_workbook(file_path, data_only=True)
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in ws.iter_rows(values_only=True):
                        # Convert row items to string
                        row_text = " ".join([str(cell) for cell in row if cell is not None])
                        content += row_text + "\n"
            
            else:
                # Default to text/binary reading as string
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()

            return IOCProcessor.extract_hashes_from_text(content)

        except Exception as e:
            raise ValueError(f"Error processing {ext} file: {e}")
